import pandas as pd
import os
import argparse
import logging
import re
from datetime import datetime
from openpyxl import Workbook
from openpyxl.cell import Cell
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

# 设置一个简单的日志器
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

def get_csv_format(df: pd.DataFrame) -> str:
    """
    根据DataFrame的列名检测CSV的格式
    """
    columns = df.columns.tolist()
    if any(col.startswith('板块') for col in columns) or any(col.startswith('标的组合') for col in columns):
        logger.info("检测到新版格式（板块-标的对）")
        return "new"
    if '价格1' in columns:
        logger.info("检测到旧版格式（含价格）")
        return "old"
    logger.warning("未能明确识别CSV格式，将按默认（新版）格式处理")
    return "new"

def auto_adjust_column_width_and_font(ws: Worksheet):
    """
    自动调整列宽以适应内容，并根据CSV格式设置字体和对齐方式
    """
    dengxian_font = Font(name='等线', size=13)
    
    col_configs = {}
    for idx, header_cell in enumerate(ws[1]):
        header = str(header_cell.value) if header_cell.value is not None else ""
        col_letter = get_column_letter(idx + 1)
        
        config = {'alignment': Alignment(), 'width': None}
        
        if header == '时间' or header.startswith('板块'):
            config['alignment'] = Alignment(horizontal='center', vertical='center')
        elif header == '标题' or header.startswith('标的组合'):
            config['alignment'] = Alignment(vertical='center', wrap_text=True)
            config['width'] = 45
        elif header in ['简述', '推荐理由', '预期']:
            config['alignment'] = Alignment(vertical='top', wrap_text=True)
            config['width'] = 45
        elif header == '原文':
            config['width'] = 666

        col_configs[col_letter] = config

    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell, Cell):
                cell.font = dengxian_font
                if cell.column_letter in col_configs:
                    cell.alignment = col_configs[cell.column_letter]['alignment']

    for col_letter, config in col_configs.items():
        if config['width']:
            ws.column_dimensions[col_letter].width = config['width']
        else:
            max_length = 0
            for cell in ws[col_letter]:
                if cell.value:
                    cell_value = str(cell.value)
                    chinese_chars = len(re.findall(r'[\u4e00-\u9fff]', cell_value))
                    other_chars = len(cell_value) - chinese_chars
                    estimated_width = chinese_chars * 2.1 + other_chars * 1.2
                    if estimated_width > max_length:
                        max_length = estimated_width
            adjusted_width = min(max(max_length + 4, 12), 60)
            ws.column_dimensions[col_letter].width = adjusted_width
            
    ws.freeze_panes = 'A2'

def process_csv_to_excel(input_csv: str, output_excel: str):
    """
    将CSV文件转换为按日期分sheet的、格式化的Excel文件
    """
    try:
        df = pd.read_csv(input_csv, encoding='utf-8-sig')
        if df.empty:
            logger.warning(f"CSV文件为空: {input_csv}")
            return
            
        logger.info(f"📊 读取CSV文件: {len(df)} 行数据")
        
        if '日期' not in df.columns:
            logger.error(f"CSV文件 {input_csv} 中缺少“日期”列，无法按日期分表。")
            return

        grouped = df.groupby('日期')
        wb = Workbook()
        if wb.active:
            wb.remove(wb.active)

        for date_group, data in sorted(grouped, key=lambda x: str(x[0]), reverse=True):
            sheet_name = str(date_group)
            ws = wb.create_sheet(title=sheet_name)
            
            data_to_write = data.copy().drop(columns=['日期'])
            
            ws.append(list(data_to_write.columns))
            for _, row in data_to_write.iterrows():
                ws.append(list(row))
            
            auto_adjust_column_width_and_font(ws)
            logger.info(f"✅ 已创建工作表: {sheet_name} ({len(data)} 行)")
            
        wb.save(output_excel)
        logger.info(f"🎉 Excel文件已生成: {output_excel}")

    except Exception as e:
        logger.error(f"❌ 转换过程中出现错误: {e}")
        raise

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="CSV to Excel 转换器，带自动格式化功能。")
    parser.add_argument('--input', type=str, default=None, help='输入的CSV文件路径。如果未提供，将自动在output/noprice_mode/daily_results/目录下查找最新的CSV文件。')
    parser.add_argument('--output', type=str, default=None, help='输出的Excel文件路径 (默认: 与输入文件同名，扩展名为.xlsx)')
    
    args = parser.parse_args()
    
    input_path = args.input
    if input_path is None:
        source_dir = 'output/noprice_mode/daily_results'
        logger.info(f"未指定输入文件，开始在 {source_dir} 目录中查找最新的CSV文件...")
        
        if not os.path.exists(source_dir):
            logger.error(f"源目录不存在: {source_dir}。请先运行main.py生成数据。")
            exit(1)

        csv_files = [os.path.join(source_dir, f) for f in os.listdir(source_dir) if f.endswith('.csv')]
        
        if not csv_files:
            logger.error(f"在 {source_dir} 目录中未找到任何CSV文件。")
            exit(1)
            
        latest_file = max(csv_files, key=os.path.getmtime)
        logger.info(f"找到最新的CSV文件: {latest_file}")
        input_path = latest_file
    
    output_path = args.output
    if output_path is None:
        if 'noprice_mode' in input_path:
             output_dir = 'output/noprice_mode/excel_reports'
        elif 'finance_mode' in input_path:
             output_dir = 'output/finance_mode/excel_reports'
        else:
             output_dir = 'output/excel_reports'
        
        os.makedirs(output_dir, exist_ok=True)
        base_name = os.path.splitext(os.path.basename(input_path))[0]
        output_path = os.path.join(output_dir, f"{base_name}.xlsx")
    
    output_dir_for_excel = os.path.dirname(output_path)
    if output_dir_for_excel and not os.path.exists(output_dir_for_excel):
        os.makedirs(output_dir_for_excel)
        
    process_csv_to_excel(input_path, output_path) 